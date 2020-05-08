import requests
import urllib.request
import time
from openpyxl import load_workbook
from bs4 import BeautifulSoup
 
def crawl(query):
    url = 'https://www.google.com/search?q='+query+'&safe=active&rlz=1C1CHBF_enIN829IN829&sxsrf=ACYBGNQPNENk2g_f3qIut60AibXoGS2b4g:1569748704119&source=lnms&tbm=nws&sa=X&ved=0ahUKEwjR4ZTP2fXkAhV47HMBHdWEA24Q_AUIEigB&biw=1366&bih=667'
    response = requests.get(url)

    soup = BeautifulSoup(response.text, 'html.parser')
    searchBlock = soup.find("id='search'")
    link = soup.find('a')
    if link is not None:
        extractedHref = link.get('href')
        if query in extractedHref:
            return True
    return False         
wb = load_workbook('test.xlsx') #File Name
ws = wb.get_sheet_by_name('Sheet1') #Sheet number


continuous_fail_count = 0
 
for row in ws.iter_rows():
    if row[1].value == "null" and continuous_fail_count < 5: # check if more than 5 sites throws error
        try:
            if crawl(row[0]._value):
                row[1].value = "Approved"
                print(row[0]._value)
                continuous_fail_count = 0
            else:
                row[1].value = "Not Approved"
                print(row[0]._value)
                continuous_fail_count = 0    
        except:
            print("An exception occurred for -- " + row[0]._value)
            continuous_fail_count = continuous_fail_count + 1
            
        
wb.save('test.xlsx') # save to file name.